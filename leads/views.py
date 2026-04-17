from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.contrib import messages
from django.db.models import Q, Count, Sum, F, Avg
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import models
from datetime import datetime, timedelta
from django.contrib.auth.models import User
from django.conf import settings as django_settings
from django.core.management import call_command
import csv
import io

from core.decorators import staff_required, superuser_required
from .models import Lead, FollowUp, CommunicationLog, Quote, CallLog, ScheduledEmail
from .source_utils import apply_source_filter, grouped_leads_by_source
from .forms import LeadForm, FollowUpForm, CommunicationLogForm, QuoteForm
try:
    from .pdf_utils import generate_and_save_quote_pdf, send_quote_pdf_email
except ImportError as e:
    # PDF functionality not available
    generate_and_save_quote_pdf = None
    send_quote_pdf_email = None
    _pdf_import_error = str(e)
from core.utils import send_email_notification
import json
import uuid

# ==================== DASHBOARD HOME ====================

@staff_required
def dashboard_home(request):
    """Dashboard homepage with comprehensive statistics and ROI metrics"""
    # Statistics
    total_leads = Lead.objects.count()
    new_leads = Lead.objects.filter(status='new').count()
    won_leads = Lead.objects.filter(status='won').count()
    lost_leads = Lead.objects.filter(status='lost').count()
    # Strict "Qualified" — same definition as Leads Management stat cards
    qualified_leads = Lead.objects.filter(status='qualified').count()
    contacted_leads = Lead.objects.filter(status__in=['contacted', 'qualified', 'proposal', 'negotiation', 'won']).count()

    new_leads_percentage = round((new_leads / total_leads * 100) if total_leads > 0 else 0, 1)
    qualified_percentage = round((qualified_leads / total_leads * 100) if total_leads > 0 else 0, 1)
    conversion_rate = round((won_leads / total_leads * 100) if total_leads > 0 else 0, 1)
    
    # Leads this week
    today = timezone.now().date()
    week_start = today - timezone.timedelta(days=today.weekday())
    new_leads_week = Lead.objects.filter(created_at__date__gte=week_start).count()
    
    # Revenue calculations
    won_quotes = Quote.objects.filter(status='accepted')
    total_revenue = won_quotes.aggregate(total=Sum('amount'))['total'] or 0
    active_customers = Quote.objects.filter(status='accepted').values('lead').distinct().count()
    avg_deal_value = (total_revenue / active_customers) if active_customers > 0 else 0
    
    leads_by_source = grouped_leads_by_source()
    
    # Leads by business type
    leads_by_business = Lead.objects.values('business_type').annotate(count=Count('business_type')).order_by('-count')
    
    # Recent leads
    recent_leads = Lead.objects.select_related('assigned_to').order_by('-created_at')[:10]
    
    # Today's follow-ups
    today_dt = timezone.now().date()
    todays_followups = FollowUp.objects.filter(
        scheduled_date__date=today_dt,
        is_completed=False
    ).select_related('lead')
    
    # Overdue follow-ups
    overdue_followups = FollowUp.objects.filter(
        scheduled_date__date__lt=today_dt,
        is_completed=False
    ).count()
    
    # Pending follow-ups (total incomplete)
    pending_followups = FollowUp.objects.filter(is_completed=False).count()
    pending_percentage = round((pending_followups / total_leads * 100) if total_leads > 0 else 0, 1)
    
    # Average response time calculations
    # Get all outbound communications ordered by lead and created_at
    all_communications = CommunicationLog.objects.filter(
        direction='out'
    ).select_related('lead').order_by('lead', 'created_at')
    
    # Deduplicate in Python to keep only the first communication per lead
    # (distinct('lead') uses PostgreSQL-specific DISTINCT ON which SQLite doesn't support)
    seen_leads = set()
    contacted_communications = []
    for comm in all_communications:
        if comm.lead_id not in seen_leads:
            contacted_communications.append(comm)
            seen_leads.add(comm.lead_id)
    
    response_times = []
    for comm in contacted_communications:
        time_diff = (comm.created_at - comm.lead.created_at).total_seconds() / 3600
        if time_diff > 0:
            response_times.append(time_diff)
            
    avg_response_time = sum(response_times) / len(response_times) if response_times else 0
    avg_response_time_display = f"{round(avg_response_time, 1)}" if response_times else "< 1"
    
    # Generate line chart data (Last 7 days)
    trend_dates = []
    trend_counts = []
    for i in range(6, -1, -1):
        target_date = today_dt - timedelta(days=i)
        count = Lead.objects.filter(created_at__date=target_date).count()
        trend_dates.append(target_date.strftime("%b %d"))
        trend_counts.append(count)
    
    context = {
        'total_leads': total_leads,
        'new_leads': new_leads,
        'won_leads': won_leads,
        'lost_leads': lost_leads,
        'qualified_leads': qualified_leads,
        'contacted_leads': contacted_leads,
        'new_leads_percentage': new_leads_percentage,
        'qualified_percentage': qualified_percentage,
        'conversion_rate': conversion_rate,
        'won_percentage': conversion_rate,
        'total_revenue': total_revenue,
        'avg_deal_value': avg_deal_value,
        'active_customers': active_customers,
        'avg_response_time': avg_response_time_display,
        'new_leads_week': new_leads_week,
        'leads_by_source': leads_by_source,
        'leads_by_business': leads_by_business,
        'recent_leads': recent_leads,
        'todays_followups': todays_followups,
        'overdue_followups': overdue_followups,
        'pending_followups': pending_followups,
        'pending_percentage': pending_percentage,
        'current_date': timezone.now(),
        'total_leads_progress': min(int(new_leads_week / 10 * 100), 100) if total_leads > 0 else 0,
        'trend_dates': trend_dates,
        'trend_counts': trend_counts,
    }
    return render(request, 'dashboard/dashboard_home.html', context)



@staff_required
def leads_list(request):
    """Display all leads with filtering"""
    leads = Lead.objects.all().select_related('assigned_to')
    
    # Filtering
    status = request.GET.get('status')
    source = request.GET.get('source')
    business_type = request.GET.get('business_type')
    
    if status:
        leads = leads.filter(status=status)
    if source:
        leads = apply_source_filter(leads, source)
    if business_type:
        leads = leads.filter(business_type=business_type)
    
    # Search
    search = request.GET.get('search')
    if search:
        leads = leads.filter(
            Q(company_name__icontains=search) |
            Q(contact_name__icontains=search) |
            Q(email__icontains=search) |
            Q(phone__icontains=search)
        )
    
    stats_total = leads.count()
    stats_new = leads.filter(status='new').count()
    stats_qualified = leads.filter(status='qualified').count()
    stats_won = leads.filter(status='won').count()

    context = {
        'leads': leads,
        'status_choices': Lead.LEAD_STATUS,
        'source_choices': Lead.LEAD_SOURCE,
        'business_types': Lead.BUSINESS_TYPES,
        'stats_total': stats_total,
        'stats_new': stats_new,
        'stats_qualified': stats_qualified,
        'stats_won': stats_won,
    }
    return render(request, 'dashboard/leads_list.html', context)


@staff_required
def add_lead(request):
    """Add a new lead"""
    if request.method == 'POST':
        # Create lead from POST data
        lead = Lead(
            company_name=request.POST.get('company_name'),
            contact_name=request.POST.get('contact_name'),
            email=request.POST.get('email'),
            phone=request.POST.get('phone'),
            business_type=request.POST.get('business_type', 'other'),
            source=request.POST.get('source', 'other'),
            service_area=request.POST.get('service_area', ''),
            address=request.POST.get('address', ''),
        )
        lead.save()
        
        messages.success(request, f'Lead "{lead.company_name}" has been added successfully.')
        return redirect('leads_list')
    
    return redirect('leads_list')


@staff_required
def edit_lead(request, pk):
    """Edit an existing lead"""
    lead = get_object_or_404(Lead, pk=pk)
    
    if request.method == 'POST':
        # Update lead from POST data
        lead.company_name = request.POST.get('company_name')
        lead.contact_name = request.POST.get('contact_name')
        lead.email = request.POST.get('email')
        lead.phone = request.POST.get('phone')
        lead.business_type = request.POST.get('business_type', 'other')
        lead.source = request.POST.get('source', 'other')
        lead.service_area = request.POST.get('service_area', '')
        lead.address = request.POST.get('address', '')
        lead.save()
        
        messages.success(request, f'Lead "{lead.company_name}" has been updated successfully.')
        return redirect('leads_list')
    
    # For GET request, return lead data as JSON for the modal
    return JsonResponse({
        'id': lead.id,
        'company_name': lead.company_name,
        'contact_name': lead.contact_name,
        'email': lead.email,
        'phone': lead.phone,
        'business_type': lead.business_type,
        'source': lead.source,
        'service_area': lead.service_area,
        'address': lead.address,
    })


@staff_required
def update_lead_status(request, pk):
    """Update lead status"""
    lead = get_object_or_404(Lead, pk=pk)
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in dict(Lead.LEAD_STATUS).keys():
            lead.status = new_status
            lead.save()
            messages.success(request, f'Lead status updated to {lead.get_status_display()}')
    return redirect('lead_detail', pk=lead.pk)


@staff_required
def follow_ups(request):
    """Display follow-ups dashboard with enhanced features"""
    from django.db.models import Q
    from .forms import FollowUpForm

    # Handle bulk completion
    if request.method == 'POST' and 'bulk_complete' in request.POST:
        selected_followups = request.POST.getlist('selected_followups')
        if selected_followups:
            FollowUp.objects.filter(pk__in=selected_followups).update(
                is_completed=True,
                completed_date=timezone.now(),
                completed_by=request.user
            )
            messages.success(request, f'Completed {len(selected_followups)} follow-up(s).')
        return redirect('follow_ups')

    # Get filter parameters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    followup_type = request.GET.get('followup_type')
    assigned_to = request.GET.get('assigned_to')
    status_filter = request.GET.get('status', 'all')  # all, upcoming, overdue, completed

    # Base queryset
    base_qs = FollowUp.objects.select_related('lead', 'completed_by')

    # Apply filters
    if date_from:
        base_qs = base_qs.filter(scheduled_date__date__gte=date_from)
    if date_to:
        base_qs = base_qs.filter(scheduled_date__date__lte=date_to)
    if followup_type:
        base_qs = base_qs.filter(followup_type=followup_type)
    if assigned_to:
        base_qs = base_qs.filter(completed_by__username=assigned_to)

    # Get statistics
    now = timezone.now()
    stats = {
        'upcoming_count': FollowUp.objects.filter(
            scheduled_date__gte=now, is_completed=False
        ).count(),
        'overdue_count': FollowUp.objects.filter(
            scheduled_date__lt=now, is_completed=False
        ).count(),
        'completed_today_count': FollowUp.objects.filter(
            completed_date__date=now.date(), is_completed=True
        ).count(),
        'total_active': FollowUp.objects.filter(is_completed=False).count(),
    }

    # Get filtered data based on status
    if status_filter == 'upcoming':
        followups = base_qs.filter(
            scheduled_date__gte=now, is_completed=False
        ).order_by('scheduled_date')
    elif status_filter == 'overdue':
        followups = base_qs.filter(
            scheduled_date__lt=now, is_completed=False
        ).order_by('scheduled_date')
    elif status_filter == 'completed':
        followups = base_qs.filter(is_completed=True).order_by('-completed_date')[:50]
    else:  # all
        # For 'all', show upcoming and overdue, plus recent completed
        upcoming = base_qs.filter(
            scheduled_date__gte=now, is_completed=False
        ).order_by('scheduled_date')
        overdue = base_qs.filter(
            scheduled_date__lt=now, is_completed=False
        ).order_by('scheduled_date')
        completed = base_qs.filter(is_completed=True).order_by('-completed_date')[:20]

        context = {
            'upcoming': upcoming,
            'overdue': overdue,
            'completed': completed,
            'stats': stats,
            'followup_form': FollowUpForm(),
            'filter_params': request.GET,
            'status_filter': status_filter,
            'leads': Lead.objects.all().order_by('company_name'),
        }
        return render(request, 'dashboard/follow_ups.html', context)

    context = {
        'followups': followups,
        'stats': stats,
        'followup_form': FollowUpForm(),
        'filter_params': request.GET,
        'status_filter': status_filter,
        'leads': Lead.objects.all().order_by('company_name'),
    }
    return render(request, 'dashboard/follow_ups_filtered.html', context)


@staff_required
def complete_followup(request, pk):
    """Mark follow-up as completed"""
    followup = get_object_or_404(FollowUp, pk=pk)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)

    followup.is_completed = True
    followup.completed_date = timezone.now()
    followup.completed_by = request.user
    followup.save()

    messages.success(request, 'Follow-up marked as completed.')

    wants_json = 'application/json' in request.headers.get('Accept', '')
    if wants_json:
        return JsonResponse({'success': True})
    return redirect('follow_ups')


@staff_required
@csrf_exempt
def create_followup(request):
    """Create a new follow-up via AJAX"""
    if request.method == 'POST':
        import json
        try:
            data = json.loads(request.body)
            lead_id = data.pop('lead', None)  # Remove lead from data and get it separately
            form = FollowUpForm(data)
            if form.is_valid() and lead_id:
                lead = get_object_or_404(Lead, pk=lead_id)
                followup = form.save(commit=False)
                followup.lead = lead
                followup.save()
                messages.success(request, 'Follow-up scheduled successfully.')
                return JsonResponse({'success': True, 'followup_id': followup.pk})
            else:
                errors = form.errors.copy()
                if not lead_id:
                    errors['lead'] = ['This field is required.']
                return JsonResponse({'success': False, 'errors': errors})
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

# ==================== PDF QUOTE GENERATION ====================

@staff_required
def create_quote(request, pk):
    """Create a new quote for lead with PDF generation"""
    lead = get_object_or_404(Lead, pk=pk)
    
    if request.method == 'POST':
        form = QuoteForm(request.POST)
        if form.is_valid():
            quote = form.save(commit=False)
            quote.lead = lead
            quote.created_by = request.user
            
            # Generate unique quote number
            quote.quote_number = f"Q-{datetime.now().strftime('%Y%m')}-{uuid.uuid4().hex[:6].upper()}"
            
            # Set default valid until (30 days from now)
            if not quote.valid_until:
                quote.valid_until = datetime.now().date() + timedelta(days=30)
            
            quote.save()
            
            # Generate PDF
            if generate_and_save_quote_pdf is None:
                messages.error(
                    request,
                    "Quote PDF generation is currently unavailable. "
                    f"({(_pdf_import_error if '_pdf_import_error' in globals() else 'unknown import error')})"
                )
                return redirect('lead_detail', pk=lead.pk)

            pdf_result = generate_and_save_quote_pdf(lead, quote)
            
            if pdf_result['success']:
                # Quote emails are sent manually by staff from the dashboard.
                # Do not auto-send on quote creation.
                messages.success(request, f"{pdf_result['message']}. Quote email is manual only.")
            else:
                messages.warning(request, pdf_result['message'])
            
            return redirect('lead_detail', pk=lead.pk)
    else:
        # Pre-populate with default values
        initial = {
            'valid_until': datetime.now().date() + timedelta(days=30),
            'terms': """Standard Terms:
1. Payment is due within 30 days of invoice date.
2. Cancellations require 24 hours notice.
3. Insurance coverage up to $2,000 per shipment.
4. Additional services billed separately.

For questions, contact your account manager."""
        }
        form = QuoteForm(initial=initial)
    
    return render(request, 'dashboard/create_quote.html', {
        'form': form,
        'lead': lead
    })


@staff_required
def view_quote_pdf(request, pk):
    """View generated PDF quote"""
    quote = get_object_or_404(Quote, pk=pk)
    
    if quote.pdf_file:
        return redirect(quote.pdf_file.url)
    else:
        messages.error(request, 'PDF file not found for this quote.')
        return redirect('lead_detail', pk=quote.lead.pk)


@staff_required
def regenerate_quote_pdf(request, pk):
    """Regenerate PDF for existing quote"""
    quote = get_object_or_404(Quote, pk=pk)
    
    if generate_and_save_quote_pdf is None:
        messages.error(
            request,
            "Quote PDF regeneration is currently unavailable. "
            f"({(_pdf_import_error if '_pdf_import_error' in globals() else 'unknown import error')})"
        )
        return redirect('lead_detail', pk=quote.lead.pk)

    pdf_result = generate_and_save_quote_pdf(quote.lead, quote)
    
    if pdf_result['success']:
        messages.success(request, 'Quote PDF regenerated successfully!')
    else:
        messages.error(request, pdf_result['message'])
    
    return redirect('lead_detail', pk=quote.lead.pk)


@staff_required
def accept_quote(request, pk):
    """Mark quote as accepted and update lead status"""
    quote = get_object_or_404(Quote, pk=pk)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)

    quote.status = 'accepted'
    quote.save()

    quote.lead.status = 'won'
    quote.lead.save()

    messages.success(request, 'Quote accepted! Lead status updated to Won.')
    return JsonResponse({'success': True})


# ==================== CALL TRACKING INTEGRATION ====================

@staff_required
def call_logs(request):
    """View all call logs"""
    calls = CallLog.objects.all().select_related('lead')
    
    context = {
        'calls': calls,
    }
    return render(request, 'dashboard/call_logs.html', context)

@staff_required
def add_call_log(request, pk):
    """Manually add a call log for a lead"""
    lead = get_object_or_404(Lead, pk=pk)
    
    if request.method == 'POST':
        from .call_tracking import CallTrackingManager
        
        call_duration = int(request.POST.get('call_duration', 0))
        call_notes = request.POST.get('call_notes', '')
        
        tracking_manager = CallTrackingManager()
        call_log = tracking_manager.log_call(
            lead=lead,
            call_duration=call_duration,
            call_notes=call_notes,
            recording_url=request.POST.get('recording_url', '')
        )
        
        # Analyze call
        tracking_manager.analyze_call(call_log)
        
        # Update lead last contacted
        lead.last_contacted = timezone.now()
        lead.save()
        
        messages.success(request, 'Call logged successfully!')
        return redirect('lead_detail', pk=lead.pk)
    
    return render(request, 'dashboard/add_call_log.html', {'lead': lead})

@csrf_exempt
def twilio_call_webhook(request):
    """Handle incoming Twilio calls for tracking"""
    if request.method == 'POST':
        token = getattr(django_settings, "TWILIO_AUTH_TOKEN", "") or ""
        if token:
            from twilio.request_validator import RequestValidator

            public_url = (django_settings.TWILIO_WEBHOOK_PUBLIC_URL or "").strip() or request.build_absolute_uri()
            signature = request.META.get("HTTP_X_TWILIO_SIGNATURE", "") or ""
            validator = RequestValidator(token)
            if not validator.validate(public_url, request.POST, signature):
                return JsonResponse({"error": "forbidden"}, status=403)

        call_data = {
            'call_sid': request.POST.get('CallSid'),
            'caller_number': request.POST.get('From'),
            'call_status': request.POST.get('CallStatus'),
            'call_duration': request.POST.get('CallDuration', 0),
            'recording_url': request.POST.get('RecordingUrl', '')
        }
        
        # Try to find lead by phone number
        from .models import Lead
        from .call_tracking import CallTrackingManager
        
        # Extract last 10 digits for matching
        caller_number = call_data['caller_number'][-10:] if call_data['caller_number'] else ''
        
        try:
            lead = Lead.objects.filter(phone__contains=caller_number).first()
        except:
            lead = None
        
        # Log the call
        tracking_manager = CallTrackingManager()
        
        # Determine call type
        call_type = 'incoming'
        if call_data['call_status'] in ['no-answer', 'busy', 'failed']:
            call_type = 'missed'
        
        # Create call log
        from .models import CallLog
        call_log = CallLog.objects.create(
            lead=lead,
            call_sid=call_data['call_sid'],
            caller_number=call_data['caller_number'],
            call_duration=int(call_data['call_duration']),
            call_notes=f"Call from {call_data['caller_number']} - Status: {call_data['call_status']}",
            recording_url=call_data['recording_url'],
            call_type=call_type
        )
        
        return JsonResponse({'status': 'success', 'call_id': call_log.id})
    
    return JsonResponse({'error': 'Invalid method'}, status=400)

# ==================== EMAIL FOLLOW-UP AUTOMATION ====================

@staff_required
def email_followup_report(request):
    """View email follow-up statistics"""
    scheduled = ScheduledEmail.objects.filter(sent=False).order_by('scheduled_date')
    sent = ScheduledEmail.objects.filter(sent=True).order_by('-sent_at')[:50]
    
    # Statistics
    total_scheduled = ScheduledEmail.objects.count()
    total_sent = ScheduledEmail.objects.filter(sent=True).count()
    open_rate = 0  # Would need tracking pixels for actual open rates
    
    context = {
        'scheduled': scheduled,
        'sent': sent,
        'total_scheduled': total_scheduled,
        'total_sent': total_sent,
        'open_rate': open_rate,
    }
    return render(request, 'dashboard/email_followups.html', context)

@staff_required
def schedule_followup_emails(request, pk):
    """Schedule follow-up emails for a lead"""
    lead = get_object_or_404(Lead, pk=pk)
    
    if request.method == 'POST':
        from .email_automation import EmailFollowUpSystem
        
        email_system = EmailFollowUpSystem(lead)
        email_system.schedule_follow_up_sequence()
        
        messages.success(request, f'Follow-up emails scheduled for {lead.company_name}')
        return redirect('lead_detail', pk=lead.pk)
    
    return render(request, 'dashboard/schedule_emails.html', {'lead': lead})

@staff_required
def send_manual_email(request, pk):
    """Manually send an email to lead"""
    lead = get_object_or_404(Lead, pk=pk)
    
    if request.method == 'POST':
        subject = request.POST.get('subject')
        message = request.POST.get('message')
        
        if subject and message:
            # Use centralized email utility
            email_sent = send_email_notification(subject, message, lead.email)
            
            if email_sent:
                # Log communication
                CommunicationLog.objects.create(
                    lead=lead,
                    communication_type='email',
                    direction='out',
                    subject=subject,
                    content=message,
                    staff_member=request.user
                )
                
                messages.success(request, f'Email sent to {lead.email}')
            else:
                messages.error(request, 'Failed to send email. Please try again.')
        else:
            messages.error(request, 'Please fill in both subject and message')
        
        return redirect('lead_detail', pk=lead.pk)
    
    return render(request, 'dashboard/send_email.html', {'lead': lead})

# ==================== UPDATED EXISTING VIEWS ====================

@staff_required
def lead_detail(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    followups = lead.followups.all()
    communications = lead.communications.all()
    quotes = lead.quotes.all()
    calls = lead.calls.all()[:10]  # Show last 10 calls
    
    if request.method == 'POST':
        if 'add_followup' in request.POST:
            form = FollowUpForm(request.POST)
            if form.is_valid():
                followup = form.save(commit=False)
                followup.lead = lead
                followup.save()
                messages.success(request, 'Follow-up scheduled successfully.')
                return redirect('lead_detail', pk=lead.pk)
        
        elif 'add_communication' in request.POST:
            form = CommunicationLogForm(request.POST)
            if form.is_valid():
                comm = form.save(commit=False)
                comm.lead = lead
                comm.staff_member = request.user
                comm.save()
                
                # Update last contacted
                lead.last_contacted = timezone.now()
                lead.save()
                
                messages.success(request, 'Communication logged successfully.')
                return redirect('lead_detail', pk=lead.pk)
    
    followup_form = FollowUpForm()
    communication_form = CommunicationLogForm()
    
    context = {
        'lead': lead,
        'followups': followups,
        'communications': communications,
        'quotes': quotes,
        'calls': calls,
        'followup_form': followup_form,
        'communication_form': communication_form,
    }
    return render(request, 'dashboard/lead_detail.html', context)

@staff_required
def export_lead_pdf(request, pk):
    """Export lead details as PDF"""
    lead = get_object_or_404(Lead, pk=pk)

    try:
        from .quote_generator import QuoteGenerator
        from .models import Quote

        temp_quote = Quote.objects.create(
            lead=lead,
            quote_number=f"INFO-{lead.pk}",
            amount=0,
            valid_until=timezone.now().date() + timedelta(days=30),
            terms="Lead information export",
            created_by=request.user
        )

        generator = QuoteGenerator(lead, temp_quote)
        pdf_filename = generator.generate_pdf_reportlab()

        if pdf_filename:
            media_url = django_settings.MEDIA_URL.rstrip('/') + '/' + pdf_filename.lstrip('/')
            return redirect(media_url)
        temp_quote.delete()
        return HttpResponse("PDF generation failed", status=500)

    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)

@staff_required
def export_leads_excel(request):
    """Export leads to Excel with custom reporting options"""
    from django.db.models import Q, F, Avg, Sum, Count
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    status_filter = request.GET.get('status')
    business_type_filter = request.GET.get('business_type')
    source_filter = request.GET.get('source')
    report_type = request.GET.get('report_type', 'standard')

    # Build queryset with filters
    leads = Lead.objects.all().order_by('-created_at')

    if start_date:
        leads = leads.filter(created_at__date__gte=start_date)
    if end_date:
        leads = leads.filter(created_at__date__lte=end_date)
    if status_filter and status_filter != 'all':
        leads = leads.filter(status=status_filter)
    if business_type_filter and business_type_filter != 'all':
        leads = leads.filter(business_type=business_type_filter)
    if source_filter and source_filter != 'all':
        leads = apply_source_filter(leads, source_filter)

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Report configuration based on type
    if report_type == 'detailed':
        headers = [
            'ID', 'Company Name', 'Contact Name', 'Email', 'Phone', 'Business Type',
            'Service Area', 'Status', 'Source', 'Created Date', 'Last Updated',
            'Notes', 'Follow-up Count', 'Quote Amount'
        ]
    elif report_type == 'analytics':
        headers = [
            'ID', 'Company Name', 'Contact Name', 'Business Type', 'Status',
            'Source', 'Created Date', 'Days Since Created', 'Conversion Status'
        ]
    else:  # standard
        headers = [
            'ID', 'Company Name', 'Contact Name', 'Email', 'Phone', 'Business Type',
            'Service Area', 'Status', 'Source', 'Created Date'
        ]

    # Set headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Add data
    for row_num, lead in enumerate(leads, 2):
        if report_type == 'detailed':
            data = [
                lead.id,
                lead.company_name,
                lead.contact_name,
                lead.email,
                lead.phone,
                lead.business_type,
                lead.service_area,
                lead.status,
                lead.source,
                lead.created_at.strftime('%Y-%m-%d %H:%M'),
                lead.updated_at.strftime('%Y-%m-%d %H:%M') if lead.updated_at else '',
                lead.notes or '',
                lead.followup_set.count(),
                lead.quote_set.filter(status='accepted').aggregate(Sum('amount'))['amount__sum'] or 0
            ]
        elif report_type == 'analytics':
            days_since_created = (timezone.now().date() - lead.created_at.date()).days
            conversion_status = 'Converted' if lead.status == 'won' else 'In Progress' if lead.status in ['contacted', 'quoted'] else 'Not Converted'
            data = [
                lead.id,
                lead.company_name,
                lead.contact_name,
                lead.business_type,
                lead.status,
                lead.source,
                lead.created_at.strftime('%Y-%m-%d'),
                days_since_created,
                conversion_status
            ]
        else:  # standard
            data = [
                lead.id,
                lead.company_name,
                lead.contact_name,
                lead.email,
                lead.phone,
                lead.business_type,
                lead.service_area,
                lead.status,
                lead.source,
                lead.created_at.strftime('%Y-%m-%d %H:%M')
            ]

        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border

    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row_num in range(1, len(leads) + 2):
            cell_value = str(ws.cell(row=row_num, column=col_num).value or '')
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Add summary sheet for analytics report
    if report_type == 'analytics' and leads.exists():
        summary_ws = wb.create_sheet("Summary")

        # Summary headers
        summary_headers = ['Metric', 'Value']
        for col_num, header in enumerate(summary_headers, 1):
            cell = summary_ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Calculate summary data
        total_leads = leads.count()
        converted_leads = leads.filter(status='won').count()
        conversion_rate = (converted_leads / total_leads * 100) if total_leads > 0 else 0
        avg_days_to_convert = leads.filter(status='won').aggregate(
            avg_days=Avg((F('updated_at') - F('created_at')).days)
        )['avg_days'] or 0

        summary_data = [
            ['Total Leads', total_leads],
            ['Converted Leads', converted_leads],
            ['Conversion Rate', f"{conversion_rate:.1f}%"],
            ['Avg Days to Convert', f"{avg_days_to_convert:.1f}"],
            ['Leads by Status', ''],
        ]

        # Add status breakdown
        status_counts = leads.values('status').annotate(count=Count('status'))
        for status_count in status_counts:
            summary_data.append([f"  {status_count['status'].title()}", status_count['count']])

        # Add summary data
        for row_num, (metric, value) in enumerate(summary_data, 2):
            summary_ws.cell(row=row_num, column=1, value=metric).border = border
            summary_ws.cell(row=row_num, column=2, value=value).border = border

        # Adjust summary column widths
        summary_ws.column_dimensions['A'].width = 25
        summary_ws.column_dimensions['B'].width = 15

    # Generate filename
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f"leads_report_{report_type}_{timestamp}.xlsx"

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response

@staff_required
def import_leads_excel(request):
    """Import leads from Excel"""
    if request.method == 'POST' and request.FILES.get('file'):
        try:
            import openpyxl
            file = request.FILES['file']
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            imported_count = 0
            errors = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    if not row[1]:  # Skip if no company name
                        continue
                    
                    lead = Lead.objects.create(
                        company_name=row[1] or '',
                        contact_name=row[2] or '',
                        email=row[3] or '',
                        phone=row[4] or '',
                        business_type=row[5] or 'other',
                        service_area=row[8] or 'GTA',
                        status='new',
                        source='import'
                    )
                    imported_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {row[0]}: {str(e)}")
            
            if imported_count > 0:
                messages.success(request, f'Successfully imported {imported_count} leads')
            if errors:
                messages.warning(request, f'Errors: {", ".join(errors[:5])}')
                
        except Exception as e:
            messages.error(request, f'Import failed: {str(e)}')
        
        return redirect('leads_list')
    
    return redirect('leads_list')

@staff_required
def dashboard_settings(request):
    """Dashboard settings"""
    from core.models import SiteSettings, Fleet, Service, EmailTemplate
    from core.service_defaults import ensure_default_services
    from django.conf import settings as django_settings
    from django.contrib.auth.models import User
    ensure_default_services()

    site_settings = SiteSettings.objects.first()
    if not site_settings:
        site_settings = SiteSettings()

    users = User.objects.all().order_by('username')
    fleet = Fleet.objects.all().order_by('name')
    fleets = Fleet.objects.all().order_by('order', 'name')
    active_fleet = Fleet.objects.filter(is_active=True).order_by('order')
    active_fleet_count = active_fleet.count()
    inactive_fleet = Fleet.objects.filter(is_active=False).order_by('name')
    services = Service.objects.all().order_by('order')

    # Email templates (fall back to defaults if not created yet)
    welcome_template = EmailTemplate.objects.filter(template_type='welcome').first()
    quote_template = EmailTemplate.objects.filter(template_type='quote').first()
    followup_template = EmailTemplate.objects.filter(template_type='followup').first()

    welcome_subject_default = f"Welcome to {django_settings.SITE_NAME}!"
    welcome_body_default = (
        "Dear {{ contact_name }},\n\n"
        f"Thank you for your interest in {django_settings.SITE_NAME}! We've received your inquiry and will get back to you shortly.\n\n"
        "In the meantime, here's what you can expect:\n"
        "- A personalized quote within 1 hour\n"
        "- Consultation on your delivery needs\n"
        "- Flexible scheduling options\n\n"
        "Best regards,\n"
        f"The {django_settings.SITE_NAME} Team"
    )

    quote_subject_default = f"Your Quote from {django_settings.SITE_NAME}"
    quote_body_default = (
        "Hello {{ contact_name }},\n\n"
        "Thanks for reaching out. Please find your quote attached: {{ quote_number }}.\n\n"
        "Valid until: {{ valid_until }}.\n\n"
        "If you have any questions, just reply to this email.\n\n"
        "Best regards,\n"
        f"The {django_settings.SITE_NAME} Team"
    )

    followup_subject_default = "Quick Follow-up: Your Delivery Quote"
    followup_body_default = (
        "Hi {{ contact_name }},\n\n"
        "Just following up regarding your delivery quote ({{ quote_number }}).\n\n"
        "Would you like to proceed, or do you have any questions?\n\n"
        "Best regards,\n"
        f"The {django_settings.SITE_NAME} Team"
    )

    if request.method == 'POST':
        if site_settings is None:
            site_settings = SiteSettings()
        site_settings.site_name = request.POST.get('company_name', site_settings.site_name)
        site_settings.site_address = request.POST.get('address', site_settings.site_address)
        site_settings.service_area = request.POST.get('service_area', site_settings.service_area)
        site_settings.business_hours = request.POST.get('business_hours', site_settings.business_hours)
        site_settings.contact_phone_1 = request.POST.get('contact_phone_1', site_settings.contact_phone_1)
        site_settings.contact_phone_2 = request.POST.get('contact_phone_2', site_settings.contact_phone_2)
        site_settings.contact_email_1 = request.POST.get('contact_email_1', site_settings.contact_email_1)
        site_settings.contact_email_2 = request.POST.get('contact_email_2', site_settings.contact_email_2)
        site_settings.sync_primary_from_contact_fields()
        site_settings.facebook_url = request.POST.get('facebook_url', site_settings.facebook_url)
        site_settings.instagram_url = request.POST.get('instagram_url', site_settings.instagram_url)
        site_settings.linkedin_url = request.POST.get('linkedin_url', site_settings.linkedin_url)
        site_settings.google_business_url = request.POST.get('google_business_url', site_settings.google_business_url)
        site_settings.twitter_url = request.POST.get('twitter_url', site_settings.twitter_url)
        site_settings.save()
        messages.success(request, 'Settings updated successfully!')
        return redirect('dashboard_settings')

    context = {
        'site_settings': site_settings,
        'users': users,
        'fleet': fleet,
        'fleets': fleets,
        'active_fleet': active_fleet,
        'active_fleet_count': active_fleet_count,
        'inactive_fleet': inactive_fleet,
        'services': services,
        'welcome_template': welcome_template,
        'quote_template': quote_template,
        'followup_template': followup_template,
        'welcome_subject_default': welcome_subject_default,
        'welcome_body_default': welcome_body_default,
        'quote_subject_default': quote_subject_default,
        'quote_body_default': quote_body_default,
        'followup_subject_default': followup_subject_default,
        'followup_body_default': followup_body_default,
        # so the template can access `{{ settings.* }}`
        'settings': django_settings,
    }
    return render(request, 'dashboard/settings.html', context)


@superuser_required
def dashboard_toggle_user_active(request, user_id):
    """Enable/disable a user from the settings page."""
    from django.contrib.auth.models import User

    if request.method != 'POST':
        return redirect('dashboard_settings')

    user = get_object_or_404(User, pk=user_id)
    # Avoid breaking access by disabling the last admin/superuser.
    if user.is_superuser and not request.POST.get('force', ''):
        messages.warning(request, "You cannot disable a superuser from this page.")
        return redirect('dashboard_settings')

    user.is_active = (request.POST.get('is_active') == 'true')
    user.save()
    messages.success(request, f"User '{user.username}' updated.")
    return redirect('dashboard_settings')


@staff_required
def dashboard_toggle_fleet_active(request, vehicle_id):
    """Enable/disable a fleet vehicle from the settings page."""
    from core.models import Fleet

    if request.method != 'POST':
        return redirect(f"{reverse('dashboard_settings')}#fleet")

    vehicle = get_object_or_404(Fleet, pk=vehicle_id)
    is_activating = (request.POST.get('is_active') == 'true')
    
    # If trying to activate, check if we can activate (max 4 active fleets)
    if is_activating:
        active_count = Fleet.objects.filter(is_active=True).exclude(pk=vehicle_id).count()
        if active_count >= 4:
            messages.error(request, "Maximum of 4 active fleet vehicles allowed. Please deactivate one before activating another.")
            return redirect(f"{reverse('dashboard_settings')}#fleet")
    
    vehicle.is_active = is_activating
    try:
        vehicle.save()
        state = "activated" if is_activating else "deactivated"
        messages.success(request, f"Vehicle '{vehicle.name}' {state}.")
    except ValueError as e:
        messages.error(request, str(e))

    return redirect(f"{reverse('dashboard_settings')}#fleet")


@staff_required
def dashboard_delete_fleet(request, vehicle_id):
    """Delete a fleet vehicle from settings page."""
    from core.models import Fleet
    from core.utils import normalize_fleet_orders

    if request.method != 'POST':
        return redirect(f"{reverse('dashboard_settings')}#fleet")

    vehicle = get_object_or_404(Fleet, pk=vehicle_id)
    vehicle_name = vehicle.name
    was_active = vehicle.is_active
    vehicle.delete()

    if was_active:
        normalize_fleet_orders()

    messages.success(request, f"Vehicle '{vehicle_name}' deleted.")
    return redirect(f"{reverse('dashboard_settings')}#fleet")


@staff_required
def dashboard_deactivate_fleet(request, vehicle_id):
    """Deactivate a fleet vehicle from the settings page. (Deprecated: use dashboard_toggle_fleet_active)"""
    from core.models import Fleet

    if request.method != 'POST':
        return redirect(f"{reverse('dashboard_settings')}#fleet")

    vehicle = get_object_or_404(Fleet, pk=vehicle_id)
    vehicle.is_active = False
    vehicle.save()
    messages.success(request, f"Vehicle '{vehicle.name}' deactivated.")
    return redirect(f"{reverse('dashboard_settings')}#fleet")


@staff_required
def dashboard_toggle_service_active(request, service_id):
    """Enable/disable a service from the settings page."""
    from core.models import Service
    from core.utils import normalize_service_orders

    if request.method != 'POST':
        return redirect('dashboard_settings')

    service = get_object_or_404(Service, pk=service_id)
    service.is_active = (request.POST.get('is_active') == 'true')
    service.save()
    
    # Normalize orders to ensure they're sequential for active services
    if service.is_active:
        normalize_service_orders()
    
    state = "activated" if service.is_active else "deactivated"
    messages.success(request, f"Service '{service.title}' {state}.")
    return redirect('dashboard_settings')


@staff_required
def dashboard_delete_service(request, service_id):
    """Delete a service from settings page."""
    from core.models import Service
    from core.utils import normalize_service_orders

    if request.method != 'POST':
        return redirect('dashboard_settings')

    service = get_object_or_404(Service, pk=service_id)
    service_title = service.title
    was_active = service.is_active
    service.delete()
    
    # Normalize orders if the deleted service was active
    if was_active:
        normalize_service_orders()
    
    messages.success(request, f"Service '{service_title}' deleted.")
    return redirect('dashboard_settings')


@staff_required
def dashboard_save_email_template(request):
    """Save email template changes from the settings page."""
    from core.models import EmailTemplate

    if request.method != 'POST':
        return redirect('dashboard_settings')

    template_type = request.POST.get('template_type')
    subject = request.POST.get('subject', '')
    body = request.POST.get('body', '')

    if template_type not in dict(EmailTemplate.TEMPLATE_TYPES):
        messages.error(request, "Invalid email template type.")
        return redirect('dashboard_settings')

    template, _ = EmailTemplate.objects.get_or_create(template_type=template_type)
    template.subject = subject
    template.body = body
    template.save()

    messages.success(request, "Email template saved successfully.")
    return redirect('dashboard_settings')


@superuser_required
def dashboard_backup_database(request):
    """Create and download a JSON backup of the database."""
    if request.method != 'GET':
        return redirect('dashboard_settings')

    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f"parcel_delivery_backup_{timestamp}.json"

    # Dump business data while excluding Django system tables that are auto-managed
    # and commonly cause UNIQUE conflicts during restore.
    buffer = io.StringIO()
    excluded_models = (
        'contenttypes',
        'auth.permission',
        'admin.logentry',
    )
    call_command(
        'dumpdata',
        *(f'--exclude={model}' for model in excluded_models),
        stdout=buffer
    )

    response = HttpResponse(buffer.getvalue(), content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


@superuser_required
def dashboard_restore_database(request):
    """Restore database from an uploaded JSON backup."""
    if request.method != 'POST' or 'backup_file' not in request.FILES:
        messages.error(request, "Please upload a backup JSON file.")
        return redirect('dashboard_settings')

    import json
    import tempfile
    import os

    backup_file = request.FILES['backup_file']

    # Create a temporary file with delete=False to avoid Windows permission issues.
    # We'll delete it manually after loading.
    tmp = tempfile.NamedTemporaryFile(suffix='.json', delete=False)
    try:
        for chunk in backup_file.chunks():
            tmp.write(chunk)
        tmp.flush()
        tmp.close()  # Close before management commands read it on Windows.

        # Sanitize uploaded fixture (works with old backups too) by removing
        # Django-managed system models that commonly cause UNIQUE conflicts.
        excluded_models = {
            'contenttypes.contenttype',
            'auth.permission',
            'admin.logentry',
        }
        with open(tmp.name, 'r', encoding='utf-8') as f:
            payload = json.load(f)
        if isinstance(payload, list):
            payload = [
                row for row in payload
                if str(row.get('model', '')).lower() not in excluded_models
            ]
        with open(tmp.name, 'w', encoding='utf-8') as f:
            json.dump(payload, f)

        # Full restore mode:
        # 1) flush existing rows
        # 2) load sanitized fixture as the source of truth
        call_command('flush', '--noinput', verbosity=0)
        call_command('loaddata', tmp.name)
        messages.success(request, "Database restored successfully.")
    except Exception as e:
        messages.error(request, f"Error restoring database: {str(e)}")
    finally:
        # Clean up the temporary file
        try:
            os.unlink(tmp.name)
        except Exception:
            pass
    
    return redirect('dashboard_settings')


@staff_required
def export_leads_csv(request):
    """Export leads to CSV (used by the settings page export button)."""
    from leads.models import Lead

    # Keep it simple: export all leads. Add filters later if needed.
    leads_qs = Lead.objects.all().order_by('id').values(
        'company_name',
        'contact_name',
        'email',
        'phone',
        'business_type',
        'source',
        'status',
        'service_area',
        'address',
    )

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([
        'company_name',
        'contact_name',
        'email',
        'phone',
        'business_type',
        'source',
        'status',
        'service_area',
        'address',
    ])
    for row in leads_qs:
        writer.writerow([row.get(col, '') for col in [
            'company_name','contact_name','email','phone',
            'business_type','source','status','service_area','address'
        ]])

    filename = f"leads_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response = HttpResponse(buffer.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response

@staff_required
def serve_logo(request):
    """Serve the logo file to prevent 404 errors"""
    from django.http import HttpResponse
    from django.conf import settings
    import os
    
    logo_path = os.path.join(settings.STATICFILES_DIRS[0], 'images', 'logo.svg')
    if os.path.exists(logo_path):
        with open(logo_path, 'rb') as f:
            return HttpResponse(f.read(), content_type='image/svg+xml')
    else:
        # Fallback to logo.png if svg not found
        logo_path = os.path.join(settings.STATICFILES_DIRS[0], 'images', 'logo.png')
        if os.path.exists(logo_path):
            with open(logo_path, 'rb') as f:
                return HttpResponse(f.read(), content_type='image/png')
        else:
            return HttpResponse(status=404)

@staff_required
def custom_reports(request):
    """Custom reports page with filtering options"""
    # Get filter options for dropdowns
    business_types = Lead.objects.values_list('business_type', flat=True).distinct()
    sources = Lead.objects.values_list('source', flat=True).distinct()
    statuses = Lead.objects.values_list('status', flat=True).distinct()

    context = {
        'business_types': sorted(business_types),
        'sources': sorted(sources),
        'statuses': sorted(statuses),
    }
    return render(request, 'dashboard/custom_reports.html', context)
@staff_required
def reports(request):
    """Reports dashboard"""
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)

    total_leads = Lead.objects.count()
    leads_this_month = Lead.objects.filter(created_at__date__gte=start_date).count()
    conversion_rate = (Lead.objects.filter(status='won').count() / total_leads * 100) if total_leads > 0 else 0

    leads_by_source = grouped_leads_by_source()

    leads_by_business = list(
        Lead.objects.values('business_type').annotate(count=Count('id')).order_by('-count')
    )
    biz_labels = dict(Lead.BUSINESS_TYPES)
    for row in leads_by_business:
        row['label'] = biz_labels.get(row['business_type'], row['business_type'])

    total_revenue = Quote.objects.filter(status='accepted').aggregate(total=Sum('amount'))['total'] or 0
    avg_deal_size = total_revenue / total_leads if total_leads > 0 else 0

    context = {
        'total_leads': total_leads,
        'leads_this_month': leads_this_month,
        'conversion_rate': round(conversion_rate, 1),
        'total_revenue': total_revenue,
        'avg_deal_size': avg_deal_size,
        'leads_by_source': leads_by_source,
        'leads_by_business': leads_by_business,
        'start_date': start_date,
        'end_date': end_date,
        'export_reports_pdf_url': reverse('export_reports_pdf'),
    }
    return render(request, 'dashboard/reports.html', context)


@staff_required
def export_reports_pdf(request):
    """Download summary analytics as PDF (matches Reports dashboard)."""
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)

    total_leads = Lead.objects.count()
    leads_this_month = Lead.objects.filter(created_at__date__gte=start_date).count()
    won = Lead.objects.filter(status='won').count()
    conversion_rate = (won / total_leads * 100) if total_leads > 0 else 0
    total_revenue = Quote.objects.filter(status='accepted').aggregate(total=Sum('amount'))['total'] or 0
    avg_deal_size = total_revenue / total_leads if total_leads > 0 else 0

    leads_by_source = grouped_leads_by_source()

    leads_by_business = list(
        Lead.objects.values('business_type').annotate(count=Count('id')).order_by('-count')
    )
    biz_labels = dict(Lead.BUSINESS_TYPES)
    for row in leads_by_business:
        row['label'] = biz_labels.get(row['business_type'], row['business_type'])

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=54,
        leftMargin=54,
        topMargin=54,
        bottomMargin=54,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=12,
    )
    story = [
        Paragraph(f"{django_settings.SITE_NAME} - Reports summary", title_style),
        Paragraph(
            f"Generated {timezone.now().strftime('%B %d, %Y at %H:%M')} - "
            f"Rolling window: {start_date} to {end_date}",
            styles['Normal'],
        ),
        Spacer(1, 16),
    ]

    summary = Table(
        [
            ['Metric', 'Value'],
            ['Total leads', str(total_leads)],
            ['New leads (30 days)', str(leads_this_month)],
            ['Conversion (won / all)', f"{conversion_rate:.1f}%"],
            ['Revenue (accepted quotes)', f"${float(total_revenue):,.2f}"],
            ['Avg. revenue per lead', f"${float(avg_deal_size):,.2f}"],
        ],
        colWidths=[3 * inch, 2.5 * inch],
    )
    summary.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ]
        )
    )
    story.append(Paragraph("<b>Key metrics</b>", styles['Heading3']))
    story.append(summary)
    story.append(Spacer(1, 20))

    src_rows = [['Source', 'Count', '% of leads']]
    for row in leads_by_source:
        pct = (row['count'] / total_leads * 100) if total_leads else 0
        src_rows.append([row['label'], str(row['count']), f"{pct:.1f}%"])

    if len(src_rows) == 1:
        src_rows.append(['No leads yet', '0', '0%'])

    src_table = Table(src_rows, colWidths=[2.8 * inch, 1 * inch, 1.2 * inch])
    src_table.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f766e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]
        )
    )
    story.append(Paragraph("<b>Leads by source</b>", styles['Heading3']))
    story.append(src_table)
    story.append(Spacer(1, 16))

    biz_rows = [['Business type', 'Count', '% of leads']]
    for row in leads_by_business:
        pct = (row['count'] / total_leads * 100) if total_leads else 0
        biz_rows.append([row['label'], str(row['count']), f"{pct:.1f}%"])

    if len(biz_rows) == 1:
        biz_rows.append(['No leads yet', '0', '0%'])

    biz_table = Table(biz_rows, colWidths=[2.8 * inch, 1 * inch, 1.2 * inch])
    biz_table.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#b45309')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]
        )
    )
    story.append(Paragraph("<b>Leads by business type</b>", styles['Heading3']))
    story.append(biz_table)

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="gta_parcel_reports_summary.pdf"'
    return response

@staff_required
def chart_data(request):
    """API endpoint to supply chart data for dashboard"""
    period = request.GET.get('period', 'week')
    today_dt = timezone.now().date()
    
    labels = []
    data = []
    
    if period == 'week':
        for i in range(6, -1, -1):
            target_date = today_dt - timedelta(days=i)
            count = Lead.objects.filter(created_at__date=target_date).count()
            labels.append(target_date.strftime("%b %d"))
            data.append(count)
    elif period == 'month':
        for i in range(29, -1, -1):
            target_date = today_dt - timedelta(days=i)
            count = Lead.objects.filter(created_at__date=target_date).count()
            labels.append(target_date.strftime("%b %d"))
            data.append(count)
    elif period == 'year':
        # Group by month for the last 12 months
        for i in range(11, -1, -1):
            target_date = today_dt.replace(day=1) - timedelta(days=i*30) 
            # This is an approximation for month stepping
            month = target_date.month
            year = target_date.year
            count = Lead.objects.filter(created_at__year=year, created_at__month=month).count()
            labels.append(target_date.strftime("%b %Y"))
            data.append(count)
            
    return JsonResponse({'labels': labels, 'data': data})

@staff_required
def request_google_review(request, pk):
    """Send an automated Google Review request email to a won lead"""
    from core.models import SiteSettings

    lead = get_object_or_404(Lead, pk=pk)
    if not (lead.email or "").strip():
        messages.error(request, "This lead has no email address to send a review request to.")
        return redirect('lead_detail', pk=lead.pk)

    ss = SiteSettings.objects.first()
    review_link = (
        (ss.google_business_url or "").strip()
        if ss
        else ""
    ) or getattr(django_settings, "SITE_GOOGLE_BUSINESS_URL", "")

    subject = f"Thank you for choosing {django_settings.SITE_NAME}!"

    context = {
        "lead": lead,
        "review_link": review_link,
        "company": {
            "name": getattr(django_settings, "SITE_NAME", "Eastern Logistics"),
            "phone": getattr(django_settings, "SITE_PHONE", ""),
            "email": getattr(django_settings, "SITE_EMAIL", ""),
            "website": getattr(django_settings, "SITE_URL", ""),
        },
    }

    try:
        from django.template.loader import render_to_string
        from django.utils.html import strip_tags

        html_message = render_to_string("email/review_request.html", context)
        message = strip_tags(html_message)
    except Exception:
        # Fallback to plaintext if template rendering fails
        message = (
            f"Hi {lead.contact_name},\n\n"
            f"Thank you for trusting {django_settings.SITE_NAME} with your logistics needs.\n\n"
            f"If you have a moment, a quick Google review helps other GTA businesses find us:\n\n"
            f"{review_link}\n\n"
            f"Best regards,\n"
            f"The {django_settings.SITE_NAME} Team\n"
        )
        html_message = None

    email_sent = send_email_notification(subject, message, lead.email, html_message=html_message)
    
    if email_sent:
        # Log communication
        CommunicationLog.objects.create(
            lead=lead,
            communication_type='email',
            direction='out',
            subject=subject,
            content=message,
            staff_member=request.user
        )
        # Update flag
        lead.google_review_requested = True
        lead.save()
        messages.success(request, f'Google Review request sent to {lead.email} successfully!')
    else:
        messages.error(request, 'Failed to send Review request email. Please check your email configuration.')
        
    return redirect('lead_detail', pk=lead.pk)